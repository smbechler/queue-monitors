"""MISO ERAS Interconnection Queue monitor.

Fetches the latest ERAS Interconnection Requests spreadsheet from MISO,
parses it, diffs it against the previous snapshot, and emails a summary.

Designed to be run on a schedule (GitHub Actions cron). The previous
snapshot is read from `snapshots/latest.json` in this directory; after
a successful run, the new snapshot is written to the same path. When
committed back to the repo, this gives a full history of the queue
over time via git.
"""

from __future__ import annotations

import json
import os
import sys
import traceback
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# Make the parent dir importable so `from lib import ...` works
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import load_workbook  # noqa: E402

from lib import diff as diff_lib  # noqa: E402
from lib import fetch, notify  # noqa: E402

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MONITOR_NAME = "MISO ERAS Queue"
LANDING_URL = (
    "https://www.misoenergy.org/planning/resource-utilization/"
    "generator-interconnection/"
)
# The xlsx filename pattern on MISO's CDN. The numeric ID and ?v= query
# parameter both change when MISO republishes, so we scrape the landing
# page for the current link rather than hardcoding it.
XLSX_PATTERN = r"ERAS[%20\s]+Interconnection[%20\s]+Requests\d+\.xlsx"

HERE = Path(__file__).resolve().parent
SNAPSHOT_DIR = HERE / "snapshots"
LATEST_SNAPSHOT = SNAPSHOT_DIR / "latest.json"
STATE_FILE_LEGACY = SNAPSHOT_DIR / "state.json"  # legacy; no longer used. Safe to delete from repo.
WORKDIR = HERE / "_workdir"

# Fields that are noisy or expected to vary; ignored when diffing rows.
# (None for now — we want to see all changes. Add fields here if needed.)
IGNORE_FIELDS: tuple[str, ...] = ()

# ---------------------------------------------------------------------------
# Fetch + parse
# ---------------------------------------------------------------------------


def find_xlsx_url() -> str:
    """Scrape the MISO landing page to find the current xlsx URL.

    Raises a clear error if the link can't be found, so the failure
    alert email tells you exactly what broke.
    """
    resp = fetch.get(LANDING_URL)
    href = fetch.find_link(resp.text, XLSX_PATTERN)
    if not href:
        raise RuntimeError(
            f"Could not find ERAS xlsx link on {LANDING_URL}. "
            f"MISO may have changed their page structure. "
            f"Update XLSX_PATTERN in {Path(__file__).name}."
        )
    # Hrefs may be relative or protocol-relative; normalize.
    if href.startswith("//"):
        href = "https:" + href
    elif href.startswith("/"):
        href = "https://www.misoenergy.org" + href
    return href


def parse_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse the ERAS spreadsheet into (projects, notes).

    The first row is the header. Each subsequent row with an
    'Application ID' becomes a project dict keyed by column name.

    Rows that have content but NO Application ID — the disclaimer and
    footnote text MISO puts below the table — are captured as `notes`:
    one string per row (non-empty cells joined), in sheet order. These
    are snapshotted and diffed like everything else, so edits to the
    fine print show up in the email too.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    headers_raw = next(rows, None)
    if not headers_raw:
        raise RuntimeError("Spreadsheet is empty — no header row found")
    headers = [_clean_header(h) for h in headers_raw]

    projects: list[dict[str, Any]] = []
    notes: list[str] = []
    for row in rows:
        if not any(cell not in (None, "") for cell in row):
            continue  # skip blank rows
        record = {headers[i]: _clean_value(row[i]) for i in range(len(headers)) if i < len(row)}
        # Use Application ID as the stable id field for diffing.
        app_id = record.get("Application ID")
        if app_id is None:
            # Not a project row — treat as sheet note/disclaimer text.
            text = " ".join(
                str(_clean_value(cell)).strip()
                for cell in row
                if cell not in (None, "")
            ).strip()
            if text:
                notes.append(text)
            continue
        record["id"] = str(app_id)
        projects.append(record)

    wb.close()
    return projects, notes


def _clean_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def _clean_value(value: Any) -> Any:
    if isinstance(value, str):
        stripped = " ".join(value.split())
        return stripped if stripped else None
    if isinstance(value, datetime):
        return value.isoformat()
    return value


# ---------------------------------------------------------------------------
# Summarize
# ---------------------------------------------------------------------------


def summarize(projects: list[dict[str, Any]]) -> dict[str, Any]:
    """Compute counts: carve-out × request-status crosstab, plus per-cycle.

    Returns:
        {
          "statuses": ["Active", "Done", "Under review", "Withdrawn", ...],
          "by_carveout": [
            {"label": "Regular",      "total": N, "by_status": {status: n}},
            {"label": "IPP",          "total": N, "by_status": {status: n}},
            {"label": "Under Review", "total": N, "by_status": {status: n}},
          ],
          "by_cycle": [
            {"label": "ERAS Study Cycle 1", "active": N, "total": N, "withdrawn": N},
            ...
          ],
        }

    Statuses are discovered from the data (alphabetical, "(blank)" last),
    so a new MISO status becomes a new column automatically.
    Carve Out matching is case-insensitive. Blank/missing values count as Regular.
    """
    # Tally projects by carve-out category. Each entry: {"total", "withdrawn"}.
    carveout_buckets: dict[str, dict[str, Any]] = {
        "Regular": {"total": 0, "by_status": {}},
        "IPP": {"total": 0, "by_status": {}},
        "Under Review": {"total": 0, "by_status": {}},
    }
    cycle_buckets: dict[str, dict[str, int]] = {}
    all_statuses: set[str] = set()

    for p in projects:
        status = _status_of(p) or "(blank)"
        all_statuses.add(status)

        # Carve-out category (case-insensitive), broken out by request status
        category = _carveout_category(p)
        carveout_buckets[category]["total"] += 1
        by_status = carveout_buckets[category]["by_status"]
        by_status[status] = by_status.get(status, 0) + 1

        # Study cycle bucket — blank cycle gets an explicit "(no cycle)" label
        # so it's visible rather than silently dropped. Each cycle also tallies
        # how many of its projects are IPP vs Regular carve-outs.
        cycle = (p.get("Study Cycle") or "").strip() or "(no cycle)"
        bucket = cycle_buckets.setdefault(
            cycle, {"total": 0, "withdrawn": 0, "ipp": 0, "regular": 0}
        )
        bucket["total"] += 1
        if _is_withdrawn(p):
            bucket["withdrawn"] += 1
        if category == "IPP":
            bucket["ipp"] += 1
        elif category == "Regular":
            bucket["regular"] += 1

    # Cycles: Active/Total/Withdrawn plus IPP/Regular carve-out counts.
    def _finalize(label: str, b: dict[str, int]) -> dict[str, Any]:
        return {
            "label": label,
            "total": b["total"],
            "withdrawn": b["withdrawn"],
            "active": b["total"] - b["withdrawn"],
            "ipp": b["ipp"],
            "regular": b["regular"],
        }

    # Status columns sorted alphabetically; (blank) goes last
    statuses = sorted(
        all_statuses,
        key=lambda s: (s == "(blank)", s.lower()),
    )

    by_carveout = [
        {
            "label": label,
            "total": carveout_buckets[label]["total"],
            "by_status": carveout_buckets[label]["by_status"],
        }
        for label in ("Regular", "IPP", "Under Review")
    ]
    # Sort cycles alphabetically; (no cycle) goes last
    cycle_labels = sorted(
        cycle_buckets.keys(),
        key=lambda c: (c == "(no cycle)", c.lower()),
    )
    by_cycle = [_finalize(c, cycle_buckets[c]) for c in cycle_labels]

    return {
        "by_carveout": by_carveout,
        "by_cycle": by_cycle,
        "statuses": statuses,
    }


def _carveout_category(project: dict[str, Any]) -> str:
    """Return one of: 'Regular', 'IPP', 'Under Review'.

    Matching is case-insensitive on the Carve Out Requested column.
    Blank/None/'N/A' all count as Regular.
    """
    raw = (project.get("Carve Out Requested") or "").strip().lower()
    if raw == "ipp":
        return "IPP"
    if raw == "under review":
        return "Under Review"
    # Everything else (blank, "N/A", or any unrecognized value) → Regular
    return "Regular"


def _status_of(project: dict[str, Any]) -> str:
    return (project.get("Request Status") or "").strip()


def _is_withdrawn(project: dict[str, Any]) -> bool:
    return (project.get("Request Status") or "").lower() == "withdrawn"


def _is_ipp(project: dict[str, Any]) -> bool:
    return (project.get("Carve Out Requested") or "").strip().lower() == "ipp"


# ---------------------------------------------------------------------------
# Snapshot I/O
# ---------------------------------------------------------------------------


def load_previous() -> tuple[list[dict[str, Any]] | None, list[str] | None]:
    """Return (projects, notes) from the last snapshot.

    `notes` is None (not []) when the snapshot predates note-tracking,
    so the first run after this upgrade doesn't misreport every
    disclaimer line as newly added.
    """
    if not LATEST_SNAPSHOT.exists():
        return None, None
    with LATEST_SNAPSHOT.open("r") as f:
        data = json.load(f)
    return data.get("projects"), data.get("notes")


def save_snapshot(
    projects: list[dict[str, Any]], xlsx_url: str, notes: list[str]
) -> None:
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    snapshot = {
        "captured_at": datetime.now(timezone.utc).isoformat(),
        "source_url": xlsx_url,
        "projects": projects,
        "notes": notes,
    }
    with LATEST_SNAPSHOT.open("w") as f:
        json.dump(snapshot, f, indent=2, default=str)


def previous_capture_time() -> str | None:
    if not LATEST_SNAPSHOT.exists():
        return None
    try:
        with LATEST_SNAPSHOT.open("r") as f:
            return json.load(f).get("captured_at")
    except Exception:  # noqa: BLE001
        return None


# ---------------------------------------------------------------------------
# Email formatting
# ---------------------------------------------------------------------------


def format_email(
    counts: dict[str, Any],
    diff: diff_lib.Diff | None,
    current: list[dict[str, Any]],
    previous: list[dict[str, Any]] | None,
    xlsx_url: str,
    last_check: str | None,
    notes: list[str],
    previous_notes: list[str] | None,
) -> tuple[str, str]:
    """Return (subject, html_body)."""
    # Notes are comparable only if the previous snapshot recorded them.
    notes_changed = previous_notes is not None and notes != previous_notes

    # Subject line packs the most important info
    total = sum(row["total"] for row in counts["by_carveout"])
    if diff is None:
        subject = f"MISO ERAS — initial snapshot ({total} projects)"
    elif diff.has_changes or notes_changed:
        bits = []
        if diff.added:
            bits.append(f"+{len(diff.added)}")
        if diff.removed:
            bits.append(f"-{len(diff.removed)}")
        if diff.changed:
            bits.append(f"~{len(diff.changed)}")
        if notes_changed:
            bits.append("notes updated")
        subject = f"MISO ERAS — {' '.join(bits)} | total {total}"
    else:
        subject = f"MISO ERAS — no changes | total {total}"

    html = _render_html(
        counts, diff, current, previous, xlsx_url, last_check,
        notes, previous_notes, notes_changed,
    )
    return subject, html


def _render_html(
    counts: dict[str, Any],
    diff: diff_lib.Diff | None,
    current: list[dict[str, Any]],
    previous: list[dict[str, Any]] | None,
    xlsx_url: str,
    last_check: str | None,
    notes: list[str],
    previous_notes: list[str] | None,
    notes_changed: bool,
) -> str:
    parts: list[str] = []
    parts.append(
        "<div style='font-family:-apple-system,BlinkMacSystemFont,Segoe UI,sans-serif;"
        "max-width:680px;color:#222;'>"
    )
    parts.append("<h2 style='margin-bottom:4px;'>MISO ERAS Queue Update</h2>")
    parts.append(
        f"<p style='color:#666;margin-top:0;font-size:13px;'>"
        f"Checked at {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        + (f" · previous check: {last_check}" if last_check else "")
        + "</p>"
    )

    # ---- Counts summary (top of email) ----
    parts.append("<h3>Summary</h3>")
    parts.append(_counts_table(counts))

    # ---- Changes ----
    parts.append("<h3 style='margin-top:24px;'>What changed</h3>")
    if diff is None:
        parts.append(
            "<p><em>This is the first snapshot — no previous data to compare against.</em></p>"
        )
    elif not diff.has_changes:
        parts.append("<p>No changes since last check.</p>")
    else:
        current_by_id = {p["id"]: p for p in current if p.get("id")}
        previous_by_id = (
            {p["id"]: p for p in previous if p.get("id")} if previous else {}
        )
        columns = _display_columns(current, previous)
        if diff.added:
            parts.append(f"<h4>Added ({len(diff.added)})</h4>")
            parts.append(_project_table(diff.added, columns))
        if diff.removed:
            parts.append(f"<h4>Removed ({len(diff.removed)})</h4>")
            parts.append(_project_table(diff.removed, columns))
        if diff.changed:
            parts.append(f"<h4>Field changes ({len(diff.changed)})</h4>")
            parts.append(
                _changed_project_table(diff.changed, current_by_id, previous_by_id, columns)
            )

    # ---- Sheet notes / disclaimers ----
    # Shown only when the footer text below the table actually changed.
    if notes_changed and previous_notes is not None:
        parts.append(
            "<h3 style='margin-top:24px;'>Sheet notes &amp; disclaimers changed</h3>"
        )
        parts.append(_notes_diff_html(previous_notes, notes))

    # ---- Source link ----
    parts.append(
        f"<p style='margin-top:24px;font-size:13px;'>"
        f"Source: <a href='{xlsx_url}'>latest spreadsheet</a> · "
        f"<a href='{LANDING_URL}'>MISO Generator Interconnection page</a>"
        f"</p>"
    )
    parts.append("</div>")
    return "".join(parts)


def _notes_diff_html(previous_notes: list[str], notes: list[str]) -> str:
    """Removed lines struck through in red, added lines in green.

    An edited line shows as one removed + one added. Duplicate lines are
    handled with multiset semantics so only genuinely changed copies show.
    """
    remaining = list(notes)
    removed: list[str] = []
    for line in previous_notes:
        if line in remaining:
            remaining.remove(line)  # unchanged copy
        else:
            removed.append(line)
    added = remaining  # current lines with no match in previous

    items: list[str] = []
    for line in removed:
        items.append(
            f"<li style='margin-bottom:6px;color:#b3261e;"
            f"text-decoration:line-through;'>{_esc(line)}</li>"
        )
    for line in added:
        items.append(
            f"<li style='margin-bottom:6px;color:#0a7d2c;font-weight:600;'>"
            f"{_esc(line)}</li>"
        )
    return (
        "<ul style='font-size:13px;padding-left:20px;list-style:none;'>"
        + "".join(items)
        + "</ul>"
    )


def _counts_table(counts: dict[str, Any]) -> str:
    """Render the summary tables: carve-out × status crosstab, then by cycle."""
    parts: list[str] = []
    parts.append(
        _carveout_status_table(
            "By Carve Out Category", counts["by_carveout"], counts["statuses"]
        )
    )
    parts.append("<div style='height:16px;'></div>")  # spacer
    parts.append(_breakdown_table("By Study Cycle", counts["by_cycle"]))
    return "".join(parts)


def _carveout_status_table(
    heading: str, rows: list[dict[str, Any]], statuses: list[str]
) -> str:
    """Crosstab: carve-out categories × request statuses.

    Columns: Total, then one per status found in the data. Status cells
    sum to Total in every row.
    """
    parts: list[str] = []
    parts.append(
        f"<h4 style='margin:8px 0 6px 0;font-size:14px;color:#444;'>"
        f"{_esc(heading)}</h4>"
    )
    parts.append("<table style='border-collapse:collapse;font-size:14px;'>")
    header_cells = [
        "<th style='text-align:left;padding:6px 12px;border-bottom:2px solid #ccc;'></th>",
        "<th style='text-align:right;padding:6px 12px;border-bottom:2px solid #ccc;'>Total</th>",
    ]
    for status in statuses:
        header_cells.append(
            f"<th style='text-align:right;padding:6px 12px;"
            f"border-bottom:2px solid #ccc;'>{_esc(status)}</th>"
        )
    parts.append("<thead><tr>" + "".join(header_cells) + "</tr></thead><tbody>")
    for row in rows:
        cells = [
            f"<td style='padding:6px 12px;font-weight:600;'>{_esc(row['label'])}</td>",
            f"<td style='padding:6px 12px;text-align:right;font-weight:600;'>{row['total']}</td>",
        ]
        for status in statuses:
            cells.append(
                f"<td style='padding:6px 12px;text-align:right;'>"
                f"{row['by_status'].get(status, 0)}</td>"
            )
        parts.append("<tr>" + "".join(cells) + "</tr>")
    parts.append("</tbody></table>")
    return "".join(parts)


def _breakdown_table(heading: str, rows: list[dict[str, Any]]) -> str:
    """Render the per-cycle table.

    Columns, in order: Active | IPP | Regular | Total | Withdrawn.
    Active = Total − Withdrawn (by request status); IPP and Regular are
    carve-out category counts within the cycle.
    """
    parts: list[str] = []
    parts.append(
        f"<h4 style='margin:8px 0 6px 0;font-size:14px;color:#444;'>"
        f"{_esc(heading)}</h4>"
    )
    parts.append("<table style='border-collapse:collapse;font-size:14px;'>")
    parts.append(
        "<thead><tr>"
        "<th style='text-align:left;padding:6px 12px;border-bottom:2px solid #ccc;'></th>"
        "<th style='text-align:right;padding:6px 12px;border-bottom:2px solid #ccc;'>Active</th>"
        "<th style='text-align:right;padding:6px 12px;border-bottom:2px solid #ccc;'>IPP</th>"
        "<th style='text-align:right;padding:6px 12px;border-bottom:2px solid #ccc;'>Regular</th>"
        "<th style='text-align:right;padding:6px 12px;border-bottom:2px solid #ccc;'>Total</th>"
        "<th style='text-align:right;padding:6px 12px;border-bottom:2px solid #ccc;'>Withdrawn</th>"
        "</tr></thead><tbody>"
    )
    for row in rows:
        parts.append(
            "<tr>"
            f"<td style='padding:6px 12px;font-weight:600;'>{_esc(row['label'])}</td>"
            f"<td style='padding:6px 12px;text-align:right;font-weight:600;'>{row['active']}</td>"
            f"<td style='padding:6px 12px;text-align:right;'>{row['ipp']}</td>"
            f"<td style='padding:6px 12px;text-align:right;'>{row['regular']}</td>"
            f"<td style='padding:6px 12px;text-align:right;'>{row['total']}</td>"
            f"<td style='padding:6px 12px;text-align:right;color:#888;'>{row['withdrawn']}</td>"
            "</tr>"
        )
    parts.append("</tbody></table>")
    return "".join(parts)


# ---------------------------------------------------------------------------
# Change tables (SPP-style: full row, with changed cells highlighted)
# ---------------------------------------------------------------------------

# The change tables show EVERY column from the spreadsheet, in the sheet's
# own column order. The column list is derived from the data at render time,
# so if MISO adds, removes, or renames a column, the email adapts with no
# code change. The internal 'id' field (a duplicate of Application ID that
# the parser adds for diffing) is the only field excluded.

_TABLE_STYLES = (
    "border-collapse:collapse;font-size:12px;font-family:-apple-system,"
    "BlinkMacSystemFont,Segoe UI,sans-serif;white-space:nowrap;"
)
_TH_STYLES = (
    "text-align:left;padding:6px 10px;background:#f4f4f4;"
    "border-bottom:2px solid #ccc;font-weight:600;"
)
_TD_STYLES = "padding:6px 10px;border-bottom:1px solid #eee;vertical-align:top;"


def _display_columns(
    current: list[dict[str, Any]],
    previous: list[dict[str, Any]] | None,
) -> list[str]:
    """Every column in the sheet, in the sheet's own order.

    Order comes from the current snapshot: parse_xlsx inserts fields in
    header order and dicts preserve insertion order. Fields that exist only
    in the previous snapshot (e.g. a column MISO renamed or dropped since
    the last run) are appended at the end so their changes still render.
    """
    columns: list[str] = []
    seen: set[str] = set()
    for project in current:
        for field in project.keys():
            if field == "id" or field in seen:
                continue
            seen.add(field)
            columns.append(field)
    if previous:
        for project in previous:
            for field in project.keys():
                if field == "id" or field in seen:
                    continue
                seen.add(field)
                columns.append(field)
    return columns


def _table_header(columns: list[str]) -> str:
    cells = "".join(
        f"<th style='{_TH_STYLES}'>{_esc(col)}</th>" for col in columns
    )
    return f"<thead><tr>{cells}</tr></thead>"


def _project_table(projects: list[dict[str, Any]], columns: list[str]) -> str:
    """Flat table for Added / Removed sections — all columns."""
    parts = ["<div style='overflow-x:auto;margin-bottom:16px;'>"]
    parts.append(f"<table style='{_TABLE_STYLES}'>")
    parts.append(_table_header(columns))
    parts.append("<tbody>")
    for project in projects:
        parts.append("<tr>")
        for field in columns:
            parts.append(
                f"<td style='{_TD_STYLES}'>{_fmt(project.get(field))}</td>"
            )
        parts.append("</tr>")
    parts.append("</tbody></table></div>")
    return "".join(parts)


def _changed_project_table(
    changes: list[tuple[str, dict[str, tuple[Any, Any]]]],
    current_by_id: dict[str, dict[str, Any]],
    previous_by_id: dict[str, dict[str, Any]],
    columns: list[str],
) -> str:
    """Table for changed projects — all columns, changed cells highlighted.

    Every cell revision the diff detected is displayed: any field in
    field_changes gets the old value struck through in red above the new
    value in green, on a yellow background.
    """
    parts = ["<div style='overflow-x:auto;margin-bottom:16px;'>"]
    parts.append(f"<table style='{_TABLE_STYLES}'>")
    parts.append(_table_header(columns))
    parts.append("<tbody>")

    for project_id, field_changes in changes:
        new_row = current_by_id.get(project_id, {})
        old_row = previous_by_id.get(project_id, {})
        parts.append("<tr>")
        for field in columns:
            if field in field_changes:
                old_val, new_val = field_changes[field]
                cell = (
                    f"<span style='color:#b3261e;text-decoration:line-through;'>"
                    f"{_fmt(old_val)}</span><br>"
                    f"<span style='color:#0a7d2c;font-weight:600;'>"
                    f"{_fmt(new_val)}</span>"
                )
                td_style = _TD_STYLES + "background:#fff8d4;white-space:normal;"
            else:
                # Unchanged cell: show current value; fall back to the old
                # value for previous-only columns (e.g. after a rename).
                value = new_row.get(field, old_row.get(field))
                cell = _fmt(value)
                td_style = _TD_STYLES
            parts.append(f"<td style='{td_style}'>{cell}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table></div>")
    return "".join(parts)


def _fmt(value: Any) -> str:
    if value is None or value == "":
        return "<em>(empty)</em>"
    return _esc(str(value))


def _esc(value: Any) -> str:
    s = str(value) if value is not None else ""
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    recipient = os.environ.get("MONITOR_TO_ADDR")
    if not recipient:
        print("ERROR: MONITOR_TO_ADDR not set", file=sys.stderr)
        return 2

    try:
        WORKDIR.mkdir(parents=True, exist_ok=True)

        print(f"[{MONITOR_NAME}] discovering current xlsx URL…")
        xlsx_url = find_xlsx_url()
        print(f"[{MONITOR_NAME}] found: {xlsx_url}")

        print(f"[{MONITOR_NAME}] downloading…")
        local_xlsx = fetch.download(xlsx_url, WORKDIR / "current.xlsx")

        print(f"[{MONITOR_NAME}] parsing…")
        projects, notes = parse_xlsx(local_xlsx)
        print(f"[{MONITOR_NAME}] parsed {len(projects)} projects, "
              f"{len(notes)} sheet note/disclaimer lines")

        counts = summarize(projects)
        print(f"[{MONITOR_NAME}] counts: {counts}")

        previous, previous_notes = load_previous()
        if previous is None:
            diff = None
        else:
            diff = diff_lib.diff_snapshots(
                previous, projects, id_field="id", ignore_fields=IGNORE_FIELDS
            )
        if previous_notes is None:
            print(f"[{MONITOR_NAME}] previous snapshot has no notes — "
                  f"note tracking starts with this run")

        last_check = previous_capture_time()
        subject, html = format_email(
            counts, diff, projects, previous, xlsx_url, last_check,
            notes, previous_notes,
        )

        print(f"[{MONITOR_NAME}] sending email: {subject}")
        notify.send_email(to=recipient, subject=subject, html=html)

        save_snapshot(projects, xlsx_url, notes)
        print(f"[{MONITOR_NAME}] snapshot saved")
        return 0

    except Exception:  # noqa: BLE001
        tb = traceback.format_exc()
        print(tb, file=sys.stderr)
        notify.send_failure_alert(recipient, MONITOR_NAME, tb)
        return 1


if __name__ == "__main__":
    sys.exit(main())
